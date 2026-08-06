import os
import re
from copy import copy
from datetime import datetime, timedelta

import openpyxl



REPORTS = [
    {
        "old_file": "REPORT_1_071726.xlsx",     
        "website_file": "website_report_1.xlsx",     
        "sheets": [
            {
                "old_sheet": "Weekly Report",           # tab in the old file
                "website_sheet": "",               # tab in the website file
                "old_header_row": 7,                    # row with column names
                "old_data_start_row": 8,                # first data row
                "website_header_row": 2,
                "website_data_start_row": 3,
                "disclaimer_marker_text": "REPLACE_WITH_DISCLAIMER_SNIPPET",
            },
 
            
        ],
    },
]

DAYS_TO_ADVANCE = 7

FILENAME_DATE_FORMAT = "%m%d%y"




def normalize(header):
    """Make header names comparable: strip, lowercase, drop 'volume_' prefix."""
    if header is None:
        return ""
    h = str(header).strip().lower()
    if h.startswith("volume_"):
        h = h[len("volume_"):]
    return h


def build_new_filename(old_path, days=DAYS_TO_ADVANCE):
    """Advance the trailing MMDDYY date in the file name by `days`."""
    folder, filename = os.path.split(old_path)
    stem, ext = os.path.splitext(filename)

    match = re.search(r"(\d{6})$", stem)
    if not match:
        raise ValueError(
            f"Could not find a trailing 6-digit MMDDYY date in '{filename}'. "
            "Expected something like REPORT_071726.xlsx"
        )

    old_date_str = match.group(1)
    old_date = datetime.strptime(old_date_str, FILENAME_DATE_FORMAT)
    new_date_str = (old_date + timedelta(days=days)).strftime(FILENAME_DATE_FORMAT)

    new_stem = stem[: match.start(1)] + new_date_str
    return os.path.join(folder, new_stem + ext)


def find_disclaimer_row(ws, start_row, marker_text):
    """Scan down from start_row; return the row index containing the marker."""
    for row in range(start_row, ws.max_row + 1):
        for cell in ws[row]:
            if cell.value and marker_text.lower() in str(cell.value).lower():
                return row
    return None


def capture_row_style(ws, row_idx):
    """Snapshot a row's values + formatting so it can be re-written elsewhere."""
    captured = []
    for cell in ws[row_idx]:
        captured.append({
            "col": cell.column,
            "value": cell.value,
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
        })
    return captured


def write_captured_row(ws, row_idx, captured):
    for info in captured:
        c = ws.cell(row=row_idx, column=info["col"], value=info["value"])
        c.font = info["font"]
        c.fill = info["fill"]
        c.border = info["border"]
        c.alignment = info["alignment"]
        c.number_format = info["number_format"]


def read_website_data(ws, header_row, data_start_row):
    headers = [normalize(c.value) for c in ws[header_row]]
    rows = []
    for row in ws.iter_rows(min_row=data_start_row):
        if all(c.value is None for c in row):
            break  # stop at the first fully blank row
        rows.append([c.value for c in row])
    return headers, rows


def reset_view_to_a1(ws):
    """Make the tab open with A1 selected / scrolled to the top-left."""
    ws.sheet_view.topLeftCell = "A1"
    try:
        ws.sheet_view.selection[0].activeCell = "A1"
        ws.sheet_view.selection[0].sqref = "A1"
    except (IndexError, AttributeError):
        pass


def process_sheet(old_wb, website_wb, cfg, old_file):
    old_sheet = cfg["old_sheet"]
    website_sheet = cfg["website_sheet"]

    if old_sheet not in old_wb.sheetnames:
        raise ValueError(f"Tab '{old_sheet}' not found in {old_file}. "
                         f"Available: {old_wb.sheetnames}")
    if website_sheet not in website_wb.sheetnames:
        raise ValueError(f"Tab '{website_sheet}' not found in the website file. "
                         f"Available: {website_wb.sheetnames}")

    old_ws = old_wb[old_sheet]
    website_ws = website_wb[website_sheet]

    print(f"  Tab '{old_sheet}'  <-  website tab '{website_sheet}'")

    # 1-2. Headers on both sides
    old_headers = [normalize(c.value) for c in old_ws[cfg["old_header_row"]]]
    website_headers, website_rows = read_website_data(
        website_ws, cfg["website_header_row"], cfg["website_data_start_row"]
    )

    if not website_rows:
        raise ValueError(f"No data rows found in website tab '{website_sheet}' - "
                         "aborting so the old data isn't wiped for nothing.")

    # 3. Map old column index -> website column index
    col_map = {}
    unmatched = []
    for old_idx, old_h in enumerate(old_headers):
        if not old_h:
            continue
        if old_h in website_headers:
            col_map[old_idx] = website_headers.index(old_h)
        else:
            unmatched.append(old_h)

    if unmatched:
        print(f"    WARNING: no website column matched: {unmatched}")
    if not col_map:
        raise ValueError(f"No columns matched between '{old_sheet}' and "
                         f"'{website_sheet}' - check the header rows in CONFIG.")

    # 4. Capture the disclaimer before wiping anything
    disclaimer_row_idx = find_disclaimer_row(
        old_ws, cfg["old_data_start_row"], cfg["disclaimer_marker_text"]
    )
    if disclaimer_row_idx is None:
        raise ValueError(f"Disclaimer not found in tab '{old_sheet}' - "
                         "check disclaimer_marker_text.")
    disclaimer_cells = capture_row_style(old_ws, disclaimer_row_idx)

    # 5. Clear old data rows + old disclaimer row
    for row in old_ws.iter_rows(min_row=cfg["old_data_start_row"],
                                max_row=old_ws.max_row):
        for cell in row:
            cell.value = None

    # 6. Write new data
    write_row = cfg["old_data_start_row"]
    for website_row in website_rows:
        for old_idx, website_idx in col_map.items():
            value = website_row[website_idx] if website_idx < len(website_row) else None
            old_ws.cell(row=write_row, column=old_idx + 1, value=value)
        write_row += 1

    last_data_row = write_row - 1

    # 7. Disclaimer goes on the row immediately after the last data row
    new_disclaimer_row = last_data_row + 1
    write_captured_row(old_ws, new_disclaimer_row, disclaimer_cells)

    reset_view_to_a1(old_ws)

    print(f"    {len(website_rows)} rows written (rows "
          f"{cfg['old_data_start_row']}-{last_data_row}); "
          f"disclaimer at row {new_disclaimer_row}")


def process_report(report):
    old_file = report["old_file"]
    website_file = report["website_file"]

    print(f"\n=== {old_file} ===")

    new_file = build_new_filename(old_file)
    if os.path.exists(new_file):
        raise ValueError(f"Output file already exists, not overwriting: {new_file}")

    old_wb = openpyxl.load_workbook(old_file)
    website_wb = openpyxl.load_workbook(website_file, data_only=True)

    for sheet_cfg in report["sheets"]:
        process_sheet(old_wb, website_wb, sheet_cfg, old_file)

    # Open the workbook on the first replaced tab, at A1
    first_tab = report["sheets"][0]["old_sheet"]
    old_wb.active = old_wb.sheetnames.index(first_tab)

    old_wb.save(new_file)
    old_wb.close()
    website_wb.close()

    print(f"  Saved as: {new_file}")


def main():
    ok, failed = 0, 0
    for report in REPORTS:
        try:
            process_report(report)
            ok += 1
        except Exception as e:
            failed += 1
            print(f"  ERROR ({report.get('old_file', '?')}): {e}")

    print(f"\nFinished. {ok} succeeded, {failed} failed.")
    input("Press Enter to close...")


if __name__ == "__main__":
    main()
